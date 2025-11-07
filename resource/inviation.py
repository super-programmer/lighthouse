import os
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment  # 用于设置超链接样式
from openpyxl.worksheet.hyperlink import Hyperlink  # 处理超链接

load_dotenv()

# 一级域名
BASE_DOMAIN = "http://zsztb.zhoushan.gov.cn"

def fetch_all_paginated_notices(url, target_date="20251030", max_pages=10):
    """分页抓取招标公告，包含项目详情链接"""
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument(f"user-agent={os.getenv('USER_AGENT', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')}")
    
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    
    all_notices = []
    current_page = 1
    target_datetime = datetime.strptime(target_date, "%Y%m%d")
    has_old_data = False

    try:
        driver.get(url)
        print(f"开始抓取第{current_page}页数据...")

        while current_page <= max_pages and not has_old_data:
            # 等待表格加载
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//table[@class='trade-table']/tbody/tr"))
                )
            except TimeoutException:
                print(f"第{current_page}页无数据，终止分页")
                break

            rows = driver.find_elements(By.XPATH, "//table[@class='trade-table']/tbody/tr")
            if not rows:
                print(f"第{current_page}页无数据，终止分页")
                break

            # 解析数据（含链接）
            page_notices = []
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) < 5:
                    continue

                # 抓取项目链接
                project_col = cols[3]
                try:
                    a_tag = project_col.find_element(By.TAG_NAME, "a")
                    relative_href = a_tag.get_attribute("href")
                    full_url = urljoin(BASE_DOMAIN, relative_href)
                except NoSuchElementException:
                    full_url = None

                notice_info = {
                    "序号": cols[0].text.strip(),
                    "所在区域": cols[1].text.strip(),
                    "公告类型": cols[2].text.strip(),
                    "项目名称": project_col.text.strip(),
                    "项目链接": full_url,
                    "发布时间": cols[4].text.strip()
                }

                # 日期筛选
                try:
                    notice_datetime = datetime.strptime(notice_info["发布时间"], "%Y-%m-%d")
                    if notice_datetime >= target_datetime:
                        page_notices.append(notice_info)
                    else:
                        has_old_data = True
                        print(f"第{current_page}页出现超期数据，终止分页")
                except ValueError:
                    print(f"日期格式异常，跳过：{notice_info['项目名称']}")
                    continue

            all_notices.extend(page_notices)
            print(f"第{current_page}页处理完成，新增{len(page_notices)}条数据")

            if has_old_data:
                break

            # 翻页
            try:
                next_page_btn = driver.find_element(
                    By.XPATH, "//a[contains(text(), '下页') or contains(text(), '»')]"
                )
                if "disabled" in next_page_btn.get_attribute("class") or not next_page_btn.is_enabled():
                    print("已到最后一页，终止分页")
                    break
                next_page_btn.click()
                current_page += 1
                WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))
            except Exception as e:
                print(f"翻页失败：{str(e)}，终止分页")
                break

        return all_notices

    finally:
        driver.quit()
        print("浏览器已关闭")

# 替换 save_to_excel_with_hyperlink 函数中的超链接设置部分
def save_to_excel_with_hyperlink(notices, filename="20251020后招标公告_带超链接.xlsx"):
    if not notices:
        print("无符合条件的招标公告，无需保存")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "招标公告"

    # 表头
    headers = ["序号", "所在区域", "公告类型", "项目名称", "项目链接", "发布时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # 填充数据（修复超链接参数）
    for row_idx, notice in enumerate(notices, 2):
        # 其他列数据（不变）
        ws.cell(row=row_idx, column=1).value = notice["序号"]
        ws.cell(row=row_idx, column=2).value = notice["所在区域"]
        ws.cell(row=row_idx, column=3).value = notice["公告类型"]
        ws.cell(row=row_idx, column=4).value = notice["项目名称"]
        ws.cell(row=row_idx, column=6).value = notice["发布时间"]

        # 修复：超链接需指定 ref 参数（单元格位置，如"E2"）
        link_cell = ws.cell(row=row_idx, column=5)  # 项目链接在第5列
        if notice["项目链接"]:
            # 生成单元格位置字符串（如第2行第5列是"E2"）
            cell_ref = f"{chr(64 + 5)}{row_idx}"  # 64是ASCII码'A'的前一位，5是第5列（E列）
            
            # 补充 ref 参数，指定单元格位置
            link_cell.hyperlink = Hyperlink(
                ref=cell_ref,  # 新增：单元格位置（关键修复）
                target=notice["项目链接"],
                display=notice["项目链接"]  # 可改为"点击查看"
            )
            link_cell.font = Font(color="0000FF", underline="single")
        else:
            link_cell.value = "无链接"

    # 调整列宽（不变）
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 12

    wb.save(filename)
    print(f"\n已保存{len(notices)}条数据到：{os.path.abspath(filename)}")

if __name__ == "__main__":
    TARGET_URL = "http://zsztb.zhoushan.gov.cn/col/col1229679799/index.html"
    TARGET_DATE = "20251020"
    MAX_PAGES = 20

    print(f"开始抓取所有分页（{TARGET_DATE}及之后），生成带超链接的Excel...")
    all_tender_notices = fetch_all_paginated_notices(TARGET_URL, TARGET_DATE, MAX_PAGES)

    if all_tender_notices:
        print(f"\n抓取完成！共获取{len(all_tender_notices)}条符合条件的招标公告")
        save_to_excel_with_hyperlink(all_tender_notices)
    else:
        print("\n未获取到符合条件的招标公告")